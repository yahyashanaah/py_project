import openpyxl as xl
from openpyxl.chart import BarChart, Reference
# Load the workbook

def proccess_workbook(filename):
    wb = xl.load_workbook(filename)

    # Access the sheet by its name
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        if isinstance(cell.value, (int, float)):  # Check if the cell value is numeric
            corrected_price = cell.value * 0.9
            corrected_price_cell = sheet.cell(row, 4)  # Create a new cell
            corrected_price_cell.value = corrected_price  # Assign the value to the new cell
        
    values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4) # Create a reference object

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')  
    # Save the workbook
    wb.save(filename)